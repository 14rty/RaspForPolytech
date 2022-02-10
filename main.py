import random

import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle


# отрывает файл для работы с ним
def OpenPlan(filename_plan):
    wb = openpyxl.load_workbook(filename_plan)
    ws = wb.active
    return ws


# функция создает карту и задаем все данные кроме предметах в семестрах
def CreateMap(filename_map):
    workbook = openpyxl.load_workbook(filename_map)
    worksheet = workbook.active
    ns = NamedStyle(name='standart')
    ns.font = Font(bold=False, size=14)
    border = Side(style='thick', color='000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    ns.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns)
    worksheet.column_dimensions['A'].height = 100
    worksheet.row_dimensions[1].height = 100
    worksheet.merge_cells('A1:I1')
    worksheet[
        'A1'] = 'КАРТА ДИСЦИПЛИН\n 09.03.03 Прикладная информатика\n Профиль "Корпоративные информационные системы", 2021 год набора, очная форма обучения'
    # worksheet['A1'] = 'КАРТА ДИСЦИПЛИН\n 09.03.03 Прикладная информатика\n Профиль "Корпоративные информационные системы", 2021 год набора, очная форма обучения'
    worksheet['A1'].style = 'standart'
    worksheet['A1'].font = Font(bold=True, size=14)
    workbook.save(filename=filename_map)
    worksheet["A2"] = "З.Е."
    worksheet['A2'].style = 'standart'
    for col in range(3, 39):
        worksheet["A" + str(2 * (col - 2) + 1)] = col - 2
        print("(=)", col - 2, "A" + str(2 * (col - 2) + 1) + ":" + "A" + str((col - 2) * 2 + 2))
        worksheet.merge_cells("A" + str(2 * (col - 2) + 1) + ":" + "A" + str((col - 2) * 2 + 2))
        worksheet["A" + str(2 * (col - 2) + 1)].style = 'standart'
    for col in range(ord('B'), ord('J')):
        worksheet[chr(col) + str(2)] = str(col - 65) + " семестр"
        worksheet[chr(col) + str(2)].style = 'standart'
    return worksheet, workbook


# def sorterBoy(filename_plan):
#     OpenPlan(filename_plan)
#     file= '\Sorter.bas'
#     os.startfile(file)

# получаем цвет пердназначенный для модуля. У каждоого модуля есть свой цвет записанный в массиве и получается посредством вызова данной функции
def add_color_cell(modules):
    colors = ['2d3eeb', '16c01d', '668f9a', 'c4b148', '5e14a8', 'd4a289', '5a6c7c', 'e0639e', 'd3fcdc', 'a7b4e7']
    colors = colors[0:len(modules)]
    s = dict()
    for i in range(len(modules)):
        print(i, " ", str(modules[i]), str(colors[i]))
        s[modules[i]] = colors[i]
    return s


#заполняем данные, размер и щвет  в ячейках карты.сначало мы загружаем данные из файла,
# потом мы сортируем данные по предметам, так чтобы одинаковыве предметы в семетрах находились на одинаковом уровне.
# Так же мы красим прелметы в соответствии с модулем
def filling_map(filename_plan, filename_map):
    ws = OpenPlan(filename_plan)
    worksheet, workbook = CreateMap(filename_map)

    zet = 0
    line = 0
    save = 0
    test = ' '
    key = 2
    term = 'B'
    modul = set()
    sem = set()
    Name = ws["D2"].value
    for i in range(2, ws.max_row + 1):
        if ws["B" + str(i)].value != None and ws["E" + str(i)].value != None:
            modul.add(ws["B" + str(i)].value)
            sem.add(ws["E" + str(i)].value)
    modul = list(modul)
    sem = list(sem)
    modul.sort()
    sem.sort()
    slow = dict()
    numer_none_slow = 1
    color_modules = add_color_cell(modul)
    print(color_modules)
    print(modul)
    for i in range(1, len(modul) + 1):
        slow[modul[i - 1]] = i
    print("-=-", slow)
    mm = []
    m = dict()
    ses = 0
    for i in range(2, ws.max_row + 1):

        if ws["I" + str(i)].value != None:

            zet = zet + ws["I" + str(i)].value

            if ws["D" + str(i)].value != ws["D" + str(i + 1)].value:
                ses = ses + 1
                Name = ws["D" + str(i)].value
                zet = int(zet)
                test = term + str(key + 1)
                cur = slow.get(ws["B" + str(i)].value)

                if cur is None:
                    cur = str(len(slow) + numer_none_slow)
                    numer_none_slow += 1
                key += zet
                dip = test + ':' + term + str(key)
                m[ses] = [Name, str(cur), zet * 2,  # todo 2 - маштаб
                          color_modules.get(str(ws["B" + str(i)].value))]
                print("+", m[ses], ws["B" + str(i)].value)

                if ws["E" + str(i)].value != ws["E" + str(i + 1)].value:
                    #     print("___")
                    mm.append(m)
                    ses = 0
                    m = dict()
                    key = 2
                    term = chr(ord(term) + 1)
        else:

            zet = 0
    mm.append(m)
    pp = []
    for x in mm:

        d = dict()
        s = 3
        p = []
        tmp = []
        for i in range(1, len(x) + 1):

            for j in range(1, len(x) + 1):
                tmp_color = x.get(j)[3]
                if (x.get(j)[3]) is None:
                    tmp_color = "" + ''.join([random.choice('0123456789ABCDEF') for j in range(6)])
                if int(x.get(j)[1]) == i:
                    if x.get(j)[0] == "Проектная деятельность" or x.get(j)[0] == "Введение в проектную деятельность":
                        tmp = [x.get(j)[0], x.get(j)[2], tmp_color]
                    else:
                        p.append([x.get(j)[0], x.get(j)[2], tmp_color])
        p.append(tmp)
        pp.append(p)
        pass

    term = 'B'
    curent = 0
    c = 0
    s = 2
    for i in range(2, ws.max_row + 1):
        if ws["I" + str(i)].value != None:

            zet = zet + ws["I" + str(i)].value

            if ws["D" + str(i)].value != ws["D" + str(i + 1)].value:
                Name = ws["D" + str(i)].value
                zet = int(zet)
                test = term + str(s + 1)
                if (len(pp[curent]) <= c):
                    pass
                s += pp[curent][c][1]
                dip = test + ':' + term + str(s)
                worksheet[test].style = 'standart'
                worksheet.merge_cells(dip)
                print(pp[curent][c], s, dip, test, str(pp[curent][c][0]))
                worksheet[test] = str(pp[curent][c][0])
                cell = worksheet[test]
                cell.fill = openpyxl.styles.PatternFill(start_color=str(pp[curent][c][2]),
                                                        end_color=str(pp[curent][c][2]),
                                                        fill_type='solid')

                c = c + 1
                if ws["E" + str(i)].value != ws["E" + str(i + 1)].value:
                    print("___")
                    curent = curent + 1
                    c = 0
                    s = 2
                    key = 2
                    term = chr(ord(term) + 1)
        else:
            zet = 0
    workbook.save(filename=filename_map)


filename_plan = '1c_Data.xlsx'
filename_map = 'map.xlsx'
wk = xlsxwriter.Workbook(filename_map)
ws = wk.add_worksheet()

ws.set_column(1, 50, 50)
wk.close()
filling_map(filename_plan, filename_map)
